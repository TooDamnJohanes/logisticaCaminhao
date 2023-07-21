import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import openpyxl


class DespachoJanela(tk.Toplevel):
    def __init__(self, frente_nome, frota_disponivel, despacho_callback, caminhao_selecionado):
        super().__init__()

        self.title(f"Frente: {frente_nome}")
        # Variável para armazenar o caminhão selecionado
        self.caminhao_selecionado = tk.StringVar()
        self.frente_nome = frente_nome  # Adicione essa linha
        self.hpc = 0  # Adicione essa linha
        self.vmc = 0  # Adicione essa linha
        self.tch = 0  # Adicione essa linha
        self.qc = 0  # Adicione essa linha
        self.nc = 0  # Adicione essa linha
        self.proximo_cam_str = ""  # Adicione essa linha
        self.raio = 0

        # Combo box para selecionar o caminhão
        ttk.Label(self, text="Caminhão:").grid(row=0, column=0, padx=10, pady=5)
        caminhao_combobox = ttk.Combobox(self, textvariable=self.caminhao_selecionado, values=frota_disponivel)
        caminhao_combobox.grid(row=0, column=1, padx=10, pady=5)

        # Botão de despachar
        despachar_button = ttk.Button(self, text="Despachar", command=self.despachar)
        despachar_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        self.despacho_callback = despacho_callback
        self.caminhao_selecionado.set(caminhao_selecionado)  # Adicione essa linha para definir o valor selecionado
        
    def despachar(self):
        caminhao = self.caminhao_selecionado.get()
        
        if caminhao:
            self.despacho_callback(caminhao)
            self.destroy()
        else:
            messagebox.showerror("Erro", "Selecione um caminhão para despachar.")


      
class FrenteColheita:
    def __init__(self, frente_frame, frota_disponivel):
        self.frente_frame = frente_frame
        self.frota_disponivel = self.buscar_frotas_caminhoes()
        self.caminhao_selecionado = tk.StringVar()  # Adicionar essa linha
        
        self.create_widgets()
        self.db_path = 'despacho.xlsx'  # Caminho do arquivo Excel
      

    def buscar_frotas_caminhoes(self):
        # Cria ou abre o arquivo Excel
        try:
            workbook = openpyxl.load_workbook('equipamentos.xlsx')
        except FileNotFoundError:
            return []

        # Seleciona a planilha
        sheet = workbook.active

        frotas = []
        # Percorre as linhas da planilha
        for row in sheet.iter_rows(min_row=0, values_only=True):
            tipo = row[1]
            frota = row[0]
            # Verifica se o tipo é "caminhão"
            if tipo == "Caminhão":
                frotas.append(frota)

        return frotas
    
    

    def create_widgets(self):
        # Campos de entrada
        self.hpc_entry = ttk.Entry(self.frente_frame)
        self.hpc_entry.grid(row=0, column=1, padx=10, pady=5)
        self.vmc_entry = ttk.Entry(self.frente_frame)
        self.vmc_entry.grid(row=1, column=1, padx=10, pady=5)
        self.tch_entry = ttk.Entry(self.frente_frame)
        self.tch_entry.grid(row=2, column=1, padx=10, pady=5)
        self.qc_entry = ttk.Entry(self.frente_frame)
        self.qc_entry.grid(row=3, column=1, padx=10, pady=5)
        self.raio_entry = ttk.Entry(self.frente_frame)
        self.raio_entry.grid(row=4, column=1, padx=10, pady=5)
        self.vcc_entry = ttk.Entry(self.frente_frame)
        self.vcc_entry.grid(row=5, column=1, padx=10, pady=5)

        # Labels
        ttk.Label(self.frente_frame, text="HPC (%):").grid(row=0, column=0, padx=10, pady=5)
        ttk.Label(self.frente_frame, text="VMC (KM/h):").grid(row=1, column=0, padx=10, pady=5)
        ttk.Label(self.frente_frame, text="TCH (Toneladas):").grid(row=2, column=0, padx=10, pady=5)
        ttk.Label(self.frente_frame, text="QC:").grid(row=3, column=0, padx=10, pady=5)
        ttk.Label(self.frente_frame, text="Raio:").grid(row=4, column=0, padx=10, pady=5)
        ttk.Label(self.frente_frame, text="VCC:").grid(row=5, column=0, padx=10, pady=5)

        # Botão Calcular
        calcular_button = ttk.Button(self.frente_frame, text="Calcular", command=self.calcular_nc)
        calcular_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

        # Labels para exibir os resultados
        self.nc_label = ttk.Label(self.frente_frame, text="NC:")
        self.nc_label.grid(row=7, column=0, padx=10, pady=5)
        self.proximo_cam_label = ttk.Label(self.frente_frame, text="Próximo envio:")
        self.proximo_cam_label.grid(row=8, column=0, padx=10, pady=5)

        # Botão de despacho
        despacho_button = ttk.Button(self.frente_frame, text="Despachar", command=self.abrir_janela_despacho)
        despacho_button.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

    def calcular_nc(self):
        
        # Obtendo os valores dos campos de entrada
        hpc = float(self.hpc_entry.get())
        vmc = float(self.vmc_entry.get())
        tch = float(self.tch_entry.get())
        qc = int(self.qc_entry.get())

        # Cálculo dos valores necessários
        hc_h = (vmc * 1.5) / 10
        tcch = hc_h * tch
        tf = tcch * qc * (hpc / 100)
        cc = 75  # Capacidade do caminhão em toneladas (exemplo)
        tcf = (cc * 60) / tf
        nc = 60 / tcf

        # Cálculo da hora do próximo envio
        now = datetime.now()
        horas_nc = int(1 / nc)
        minutos_extra = int((1 / nc - horas_nc) * 60)
        proximo_cam = now + timedelta(hours=horas_nc, minutes=minutos_extra)
        proximo_cam_str = proximo_cam.strftime("%Y-%m-%d %H:%M:%S")

  
        # Atualiza os valores na interface
        self.nc_label.configure(text=f"NC: {nc:.2f}")
        self.proximo_cam_label.configure(text=f"Próximo envio: {proximo_cam_str}")

        
         # Cálculo do ciclo
        vcc = float(self.vcc_entry.get())
        raio = float(self.raio_entry.get())
        ciclo = float(raio / vcc)
        hora_envio = datetime.now()
        # Cálculo da hora de retorno
        hora_retorno = hora_envio + timedelta(hours=ciclo)
        hora_retorno_str = hora_retorno.strftime("%Y-%m-%d %H:%M:%S")
        
         
    def abrir_janela_despacho(self):
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']
        self.hpc = float(self.hpc_entry.get()) # linhas para armazenar os valores
        self.vmc = float(self.vmc_entry.get())
        self.tch = float(self.tch_entry.get())
        self.qc = int(self.qc_entry.get())
        self.nc = float(self.nc_label["text"].split(": ")[1])
        self.proximo_cam_str = self.proximo_cam_label["text"].split(": ")[1]
        caminhao = self.caminhao_selecionado.get()
        self.raio = float(self.raio_entry.get())
        self.vcc = float(self.vcc_entry.get())
        #hora_retorno_c = self.hora_retorno_str

       
        janela_despacho = DespachoJanela(frente_nome, self.frota_disponivel, self.despacho_caminhao, caminhao)

        janela_despacho.grab_set()


       

    def despacho_caminhao(self, caminhao):
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']
        hpc = self.hpc  # Obtenha os valores armazenados
        vmc = self.vmc
        tch = self.tch
        qc = self.qc
        nc = self.nc
        proximo_cam_str = self.proximo_cam_str
        raio = self.raio
        #hora_retorno_c = self.hora_retorno_str
        #self.buscar_hora_envio(hr_retorno)
        self.inserir_dados_frente(hpc, vmc, tch, qc, nc, proximo_cam_str, caminhao, raio)
        
    
    def inserir_dados_frente(self, hpc, vmc, tch, qc, nc, proximo_cam, caminhao, raio, hora_retorno_c):
        # Cria ou abre o arquivo Excel
        try:
            workbook = openpyxl.load_workbook(self.db_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.save(self.db_path)

        # Seleciona a planilha
        sheet = workbook.active

        # Verifica se a planilha está vazia
        if sheet.max_row == 1:
            # Insere os cabeçalhos das colunas
            sheet.cell(row=1, column=1, value="HPC (%)")
            sheet.cell(row=1, column=2, value="VMC (KM/h)")
            sheet.cell(row=1, column=3, value="TCH (Toneladas)")
            sheet.cell(row=1, column=4, value="QC")
            sheet.cell(row=1, column=5, value="NC")
            sheet.cell(row=1, column=6, value="Prox_envio")
            sheet.cell(row=1, column=7, value="Data e Hora")
            sheet.cell(row=1, column=8, value="Nome da Frente")
            sheet.cell(row=1, column=9, value="Caminhão")
            sheet.cell(row=1, column=10, value="Raio")
            #sheet.cell(row=1, column=11, value="Hora retorno")

        # Obtém a última linha preenchida na planilha
        last_row = sheet.max_row

        # Obtém o nome da frente a partir do título da aba
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']

        # Insere os dados na próxima linha vazia
        sheet.cell(row=last_row + 1, column=1, value=hpc)
        sheet.cell(row=last_row + 1, column=2, value=vmc)
        sheet.cell(row=last_row + 1, column=3, value=tch)
        sheet.cell(row=last_row + 1, column=4, value=qc)
        sheet.cell(row=last_row + 1, column=5, value=nc)
        sheet.cell(row=last_row + 1, column=6, value=proximo_cam)
        sheet.cell(row=last_row + 1, column=7, value=datetime.now())
        sheet.cell(row=last_row + 1, column=8, value=frente_nome)
        sheet.cell(row=last_row + 1, column=9, value=caminhao)
        sheet.cell(row=last_row + 1, column=10, value=raio)
       # sheet.cell(row=last_row + 1, column=11, value=hora_retorno_c)
        

        # Salva as alterações no arquivo Excel
        workbook.save(self.db_path)


# Criação da janela principal
janela_principal = tk.Tk()
janela_principal.title("Sistema de Despacho de Caminhões")

# Lista de frotas disponíveis
frota_disponivel = ['Frota 1', 'Frota 2', 'Frota 3']

# Criação das abas para as frentes de colheita
abas_frentes = ttk.Notebook(janela_principal)

# Aba para a primeira frente de colheita
frente1_frame = ttk.Frame(abas_frentes)
abas_frentes.add(frente1_frame, text="Frente 1")
frente1 = FrenteColheita(frente1_frame, frota_disponivel)

# Aba para a segunda frente de colheita
frente2_frame = ttk.Frame(abas_frentes)
abas_frentes.add(frente2_frame, text="Frente 2")
frente2 = FrenteColheita(frente2_frame, frota_disponivel)

# Adicione mais abas de acordo com suas necessidades

# Posicionamento das abas
abas_frentes.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

# Execução da janela principal
janela_principal.mainloop()