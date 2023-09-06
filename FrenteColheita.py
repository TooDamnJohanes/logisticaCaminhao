import tkinter as tk
from datetime import datetime as date_time

from datetime import timedelta
from tkinter import ttk

from openpyxl.utils import datetime

from Constants import CAMINHAO_PATH_FOLDER_NAME
from Despacho import Despacho
from GerenciaArquivosExcel import GerenciaArquivosExcel


class FrenteColheita:
    DESPACHO_PLANILHA_FILE_NAME = 'despacho.xlsx'
    CAMINHAO = "Caminhão"
    HPC_LABEL_TEXT = "HPC (%):"
    VMC_LABEL_TEXT = "VMC (KM/h):"
    TCH_LABEL_TEXT = "TCH (Toneladas):"
    QC_LABEL_TEXT = "QC:"
    RAIO_LABEL_TEXT = "Raio:"
    VCC_LABEL_TEXT = "VCC:"
    CALCULAR = "Calcular"
    NC = "NC:"
    PROXIMO_ENVIO_LABEL = "Próximo envio:"
    DESPACHAR = "Despachar"

    def __init__(self, frente_frame):
        self.qc_entry = None
        self.tch_entry = None
        self.vmc_entry = None
        self.hpc_entry = None
        self.raio_entry = None
        self.frente_frame = frente_frame
        self.gerenciador_planilha_despacho = GerenciaArquivosExcel(
            path_arquivo_ser_criado=CAMINHAO_PATH_FOLDER_NAME,
            nome_planilha=self.DESPACHO_PLANILHA_FILE_NAME
        )
        self.frota_disponivel = self.buscar_frotas_caminhoes()
        self.caminhao_selecionado = tk.StringVar()  # Adicionar essa linha
        self.create_widgets()

    def buscar_frotas_caminhoes(self):
        sheet = self.gerenciador_planilha_despacho.cria_arquivo_excel().active
        frotas = []
        # Percorre as linhas da planilha
        for row in sheet.iter_rows(min_row=0, values_only=True):
            tipo = row[1]
            frota = row[0]
            # Verifica se o tipo é "caminhão"
            if tipo == self.CAMINHAO:
                frotas.append(frota)

        return frotas

    def cria_hpc_entry_component(self):
        self.hpc_entry = ttk.Entry(self.frente_frame)
        self.hpc_entry.grid(row=0, column=1, padx=10, pady=5)

    def cria_vmc_entry_component(self):
        self.vmc_entry = ttk.Entry(self.frente_frame)
        self.vmc_entry.grid(row=1, column=1, padx=10, pady=5)

    def cria_tch_entry_component(self):
        self.tch_entry = ttk.Entry(self.frente_frame)
        self.tch_entry.grid(row=2, column=1, padx=10, pady=5)

    def cria_qc_entry_component(self):
        self.qc_entry = ttk.Entry(self.frente_frame)
        self.qc_entry.grid(row=3, column=1, padx=10, pady=5)

    def cria_raio_entry_component(self):
        self.raio_entry = ttk.Entry(self.frente_frame)
        self.raio_entry.grid(row=4, column=1, padx=10, pady=5)

    def cria_vcc_entry_component(self):
        self.vcc_entry = ttk.Entry(self.frente_frame)
        self.vcc_entry.grid(row=5, column=1, padx=10, pady=5)

    def cria_hpc_label_component(self):
        ttk.Label(self.frente_frame, text=self.HPC_LABEL_TEXT).grid(row=0, column=0, padx=10, pady=5)

    def cria_vmc_label_component(self):
        ttk.Label(self.frente_frame, text="VMC (KM/h):").grid(row=1, column=0, padx=10, pady=5)

    def cria_tch_label_component(self):
        ttk.Label(self.frente_frame, text="TCH (Toneladas):").grid(row=2, column=0, padx=10, pady=5)

    def cria_qc_label_component(self):
        ttk.Label(self.frente_frame, text="QC:").grid(row=3, column=0, padx=10, pady=5)

    def cria_raio_label_component(self):
        ttk.Label(self.frente_frame, text="Raio:").grid(row=4, column=0, padx=10, pady=5)

    def cria_vcc_label_component(self):
        ttk.Label(self.frente_frame, text="VCC:").grid(row=5, column=0, padx=10, pady=5)

    def cria_nc_label_component(self):
        self.nc_label = ttk.Label(self.frente_frame, text=self.NC)
        self.nc_label.grid(row=7, column=0, padx=10, pady=5)

    def cria_proximo_caminhao_label(self):
        self.proximo_cam_label = ttk.Label(self.frente_frame, text=self.PROXIMO_ENVIO_LABEL)
        self.proximo_cam_label.grid(row=8, column=0, padx=10, pady=5)

    def cria_label_components(self):
        self.cria_hpc_label_component()
        self.cria_vmc_label_component()
        self.cria_tch_label_component()
        self.cria_qc_label_component()
        self.cria_raio_label_component()
        self.cria_vcc_label_component()
        self.cria_nc_label_component()
        self.cria_proximo_caminhao_label()

    def cria_entry_components(self):
        self.cria_hpc_entry_component()
        self.cria_vmc_entry_component()
        self.cria_tch_entry_component()
        self.cria_qc_entry_component()
        self.cria_raio_entry_component()
        self.cria_vcc_entry_component()

    def cria_botao_calcular(self):
        calcular_button = ttk.Button(self.frente_frame, text=self.CALCULAR, command=self.calcular_nc)
        calcular_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

    def cria_botao_despacho(self):
        despacho_button = ttk.Button(self.frente_frame, text=self.DESPACHAR, command=self.abrir_janela_despacho)
        despacho_button.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

    def create_widgets(self):
        # Campos de entrada
        self.cria_entry_components()

        # Labels
        self.cria_label_components()

        # Botão Calcular
        self.cria_botao_calcular()

        # Botão de despacho
        self.cria_botao_despacho()

    def calcular_nc(self):
        nc = self.calcula_nc_value()

        # Cálculo da hora do próximo envio
        proximo_cam_str = (self.calcula_horario_proximo_envio(nc=nc)
                           .strftime("%Y-%m-%d %H:%M:%S"))

        # Atualiza os valores na interface
        self.nc_label.configure(text=f"NC: {nc:.2f}")
        self.proximo_cam_label.configure(text=f"Próximo envio: {proximo_cam_str}")

        # Cálculo do ciclo
        ciclo = self.calcula_horario_ciclo()
        hora_envio = date_time.now()
        # Cálculo da hora de retorno
        hora_retorno = hora_envio + timedelta(hours=ciclo)
        hora_retorno_str = hora_retorno.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def calcula_hc_h_value(vmc):
        return (vmc * 1.5) / 10

    @staticmethod
    def calcula_tcch_value(hc_h, tch):
        return hc_h * tch

    @staticmethod
    def calcula_tf_value(tcch, qc, hpc):
        return tcch * qc * (hpc / 100)

    @staticmethod
    def calcula_tcf_value(cc, tf):
        return (cc * 60) / tf

    def calcula_nc_value(self):
        # Obtendo os valores dos campos de entrada
        hpc = float(self.hpc_entry.get())
        vmc = float(self.vmc_entry.get())
        tch = float(self.tch_entry.get())
        qc = int(self.qc_entry.get())

        # Cálculo dos valores necessários
        hc_h = self.calcula_hc_h_value(vmc=vmc)
        tcch = self.calcula_tcch_value(hc_h=hc_h, tch=tch)
        tf = self.calcula_tf_value(
            tcch=tcch,
            qc=qc,
            hpc=hpc
        )
        cc = 75  # Capacidade do caminhão em toneladas (exemplo)
        return 60 / self.calcula_tcf_value(cc=cc, tf=tf)

    @staticmethod
    def calcula_horario_proximo_envio(nc):
        now = date_time.now()
        horas_nc = int(1 / nc)
        minutos_extra = int((1 / nc - horas_nc) * 60)
        return now + timedelta(hours=horas_nc, minutes=minutos_extra)

    def calcula_horario_ciclo(self):
        vcc = float(self.vcc_entry.get())
        raio = float(self.raio_entry.get())
        return float(raio / vcc)

    def abrir_janela_despacho(self):
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']
        self.hpc = float(self.hpc_entry.get())  # linhas para armazenar os valores
        self.vmc = float(self.vmc_entry.get())
        self.tch = float(self.tch_entry.get())
        self.qc = int(self.qc_entry.get())
        self.nc = float(self.nc_label["text"].split(": ")[1])
        self.proximo_cam_str = self.proximo_cam_label["text"].split(": ")[1]
        self.raio = float(self.raio_entry.get())
        self.vcc = float(self.vcc_entry.get())

        janela_despacho = Despacho(frente_nome, self.frota_disponivel, self.despacho_caminhao,
                                   self.caminhao_selecionado.get())

        janela_despacho.grab_set()

    def despacho_caminhao(self, caminhao):
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']
        hpc = self.hpc  # Obtenha os valores armazenados
        vmc = self.vmc
        tch = self.tch
        qc = self.qc
        nc = self.nc
        proximo_cam_str = self.proximo_cam_str
        raio = self.raio
        # hora_retorno_c = self.hora_retorno_str
        # self.buscar_hora_envio(hr_retorno)
        self.inserir_dados_frente(hpc, vmc, tch, qc, nc, proximo_cam_str, caminhao, raio)

    def inserir_dados_frente(self, hpc, vmc, tch, qc, nc, proximo_cam, caminhao, raio):
        # Seleciona a planilha
        sheet = self.gerenciador_planilha_despacho.seleciona_planilha_excel()

        # Verifica se a planilha está vazia
        if sheet.max_row == 1:
            # Insere os cabeçalhos das colunas
            sheet.cell(row=1, column=1, value="HPC (%)")
            sheet.cell(row=1, column=2, value="VMC (KM/h)")
            sheet.cell(row=1, column=3, value="TCH (Toneladas)")
            sheet.cell(row=1, column=4, value="QC")
            sheet.cell(row=1, column=5, value="NC")
            sheet.cell(row=1, column=6, value="Prox_envio")
            sheet.cell(row=1, column=7, value="Data e Hora")
            sheet.cell(row=1, column=8, value="Nome da Frente")
            sheet.cell(row=1, column=9, value="Caminhão")
            sheet.cell(row=1, column=10, value="Raio")
            # sheet.cell(row=1, column=11, value="Hora retorno")

        # Obtém a última linha preenchida na planilha
        last_row = sheet.max_row

        # Obtém o nome da frente a partir do título da aba
        frente_nome = self.frente_frame.master.tab(self.frente_frame)['text']

        # Insere os dados na próxima linha vazia
        sheet.cell(row=last_row + 1, column=1, value=hpc)
        sheet.cell(row=last_row + 1, column=2, value=vmc)
        sheet.cell(row=last_row + 1, column=3, value=tch)
        sheet.cell(row=last_row + 1, column=4, value=qc)
        sheet.cell(row=last_row + 1, column=5, value=nc)
        sheet.cell(row=last_row + 1, column=6, value=proximo_cam)
        sheet.cell(row=last_row + 1, column=7, value=datetime.now())
        sheet.cell(row=last_row + 1, column=8, value=frente_nome)
        sheet.cell(row=last_row + 1, column=9, value=caminhao)
        sheet.cell(row=last_row + 1, column=10, value=raio)
        # sheet.cell(row=last_row + 1, column=11, value=hora_retorno_c)

        # Salva as alterações no arquivo Excel
        self.gerenciador_planilha_despacho.salva_planilha()
