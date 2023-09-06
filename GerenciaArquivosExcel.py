import openpyxl


class GerenciaArquivosExcel:

    def __init__(self, path_arquivo_ser_criado, nome_planilha):
        self.path_arquivo_ser_criado = path_arquivo_ser_criado
        self.nome_planilha = nome_planilha
        self.work_book = self.cria_arquivo_excel()
        self.planilha_atual = self.seleciona_planilha_excel()

    def cria_arquivo_excel(self):
        try:
            return openpyxl.load_workbook(self.path_arquivo_ser_criado)
        except FileNotFoundError:
            return openpyxl.Workbook()

    def seleciona_planilha_excel(self):
        if self.nome_planilha in self.work_book.sheetnames:
            return self.work_book.get_sheet_by_name(self.nome_planilha)
        else:
            return self.work_book.create_sheet(self.nome_planilha)

    def salva_planilha(self):
        self.work_book.save(self.path_arquivo_ser_criado)

    def insere_cabecalho_planilha(self, colum, value):
        if self.planilha_atual.max_row == 1:
            # Insere os cabeçalhos das colunas
            self.planilha_atual.cell(row=1, column=colum, value=value)

    def insere_dados_proxima_linha(self, row, column, value):
        self.planilha_atual.cell(row=row, column=column, value=value)

    def retorna_ultima_linha_planilha(self):
        return self.planilha_atual.max_row
