import tkinter as tk
from tkinter import ttk
from tkinter import messagebox  # Importe o módulo messagebox separadamente
import openpyxl

from Constants import CAMINHAO_PATH_FOLDER_NAME


class CadastroEquipamentos:
    EQUIPAMENTOS = 'equipamentos'
    FROTA = "Frota"
    TIPO = "Tipo"
    MODELO = "Modelo"
    GRUPO = "Grupo"
    MESSAGEBOX_DIALOG_TITLE = "Cadastro de Equipamentos"
    MESSAGEBOX_DIALOG_DESCRIPTION = "Equipamento cadastrado com sucesso!"

    def __init__(self):
        # Campos de entrada
        self.frota_entry = ttk.Entry()
        self.tipo_entry = ttk.Entry()
        self.modelo_entry = ttk.Entry()
        self.grupo_entry = ttk.Entry()
        self.db_path = CAMINHAO_PATH_FOLDER_NAME

    def cadastrar_equipamento(self):
        frota = self.frota_entry.get()
        tipo = self.tipo_entry.get()
        modelo = self.modelo_entry.get()
        grupo = self.grupo_entry.get()

        # Cria ou abre o arquivo Excel
        workbook = self.cria_arquivo_excel()

        # Seleciona a planilha "equipamentos" ou cria uma nova
        sheet = self.seleciona_planilha_excel(workbook=workbook)

        # Verifica se a planilha está vazia
        self.verifica_planilha_vazia(sheet=sheet)

        # Insere os dados na próxima linha vazia
        self.insere_dados_proxima_linha(
            sheet=sheet,
            frota=frota,
            tipo=tipo,
            modelo=modelo,
            grupo=grupo
        )
        # Salva as alterações no arquivo Excel
        self.salva_planilha(workbook=workbook)

        # Limpa os campos de entrada após o cadastro
        self.limpa_campos_apos_cadastro()

        # Exibe uma mensagem de sucesso
        self.dialog_equipamento_cadastrado()

    def cria_arquivo_excel(self):
        try:
            return openpyxl.load_workbook(self.db_path)
        except FileNotFoundError:
            return openpyxl.Workbook()

    def seleciona_planilha_excel(self, workbook):
        if self.EQUIPAMENTOS in workbook.sheetnames:
            return workbook.get_sheet_by_name(self.EQUIPAMENTOS)
        else:
            return workbook.create_sheet(self.EQUIPAMENTOS)

    def verifica_planilha_vazia(self, sheet):
        if sheet.max_row == 1:
            # Insere os cabeçalhos das colunas
            sheet.cell(row=1, column=1, value=self.FROTA)
            sheet.cell(row=1, column=2, value=self.TIPO)
            sheet.cell(row=1, column=3, value=self.MODELO)
            sheet.cell(row=1, column=4, value=self.GRUPO)

    def insere_dados_proxima_linha(self, sheet, frota, tipo, modelo, grupo):
        last_row = sheet.max_row
        sheet.cell(row=last_row + 1, column=1, value=frota)
        sheet.cell(row=last_row + 1, column=2, value=tipo)
        sheet.cell(row=last_row + 1, column=3, value=modelo)
        sheet.cell(row=last_row + 1, column=4, value=grupo)

    def salva_planilha(self, workbook):
        workbook.save(self.db_path)

    def limpa_campos_apos_cadastro(self):
        self.frota_entry.delete(0, tk.END)
        self.tipo_entry.delete(0, tk.END)
        self.modelo_entry.delete(0, tk.END)
        self.grupo_entry.delete(0, tk.END)

    def dialog_equipamento_cadastrado(self):
        messagebox.showinfo(self.MESSAGEBOX_DIALOG_TITLE, self.MESSAGEBOX_DIALOG_DESCRIPTION)
